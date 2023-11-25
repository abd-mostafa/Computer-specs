from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
import openpyxl

import os

from openpyxl import Workbook, load_workbook



# to load the workbook with its path
wb = openpyxl.load_workbook('ST.xlsx')
s = wb.active


class specs():
  def pc_specs(self):
    for i in range(2, s.max_row + 1):
         driver = webdriver.Edge(executable_path = EdgeChromiumDriverManager().install())
         driver.get("https://www.dell.com/support/home/en-us")
         element = driver.find_element(By.XPATH, "//input[@id='mh-search-input']")
         element.send_keys(s.cell (row=i, column=1).value)
         # WebDriverWait(driver, 20).until(
         #     EC.visibility_of_element_located(By.CSS_SELECTOR, "button[aria-label='Rechercher dans le support Dell']"))
         # click_search = driver.find_element(By.XPATH, "//button[@aria-label='Rechercher dans le support Dell']")
         # click_search.click()
         WebDriverWait(driver, 20).until(
             EC.element_to_be_clickable((By.XPATH, '//*[@id="unified-masthead"]/div[1]/div[1]/div[2]/button[2]'))).click()
         try:
          model = WebDriverWait(driver, 20).until(
          EC.visibility_of_element_located((By.CSS_SELECTOR , "h1[aria-label='SystemDescription']"))).text
          s.cell(row=i, column=2).value = model
          print(model)
          EXP_DATE = WebDriverWait(driver, 20).until(
             EC.visibility_of_element_located((By.CSS_SELECTOR, ".warrantyExpiringLabel.mb-0.ml-1.mr-1"))).text
          s.cell(row=i, column=4).value = EXP_DATE
          print(EXP_DATE)
          viewproductspecs = WebDriverWait(driver, 20).until(
             EC.visibility_of_element_located((By.XPATH, "//a[@id='quicklink-sysconfig']")))
          viewproductspecs.click()
         # all_spans = WebDriverWait(driver, 20).until(
         #     EC.presence_of_all_elements_located((By.CLASS_NAME, "font-weight-medium text-jet pr-4")))
         # all_spans = driver.find_elements(By.CLASS_NAME, "font-weight-medium text-jet pr-4")
         # for span in all_spans:
         #     print(span.text)
          all_spans = WebDriverWait(driver, 20).until(
             EC.presence_of_all_elements_located((By.XPATH, "//span[@class='font-weight-medium text-jet pr-4']")))
          for span in all_spans:
            if 'Core i5' in span.text:
              s.cell(row=i, column=3).value = 'Core i5'
              print(span.text)
              wb.save('ST.xlsx')
              driver.close()
              break
            elif 'Core i7'in span.text:
              s.cell(row=i, column=3).value = 'Core i7'
              print(span.text)
              wb.save('ST.xlsx')
              driver.close()
              break
            elif 'Core i3'in span.text:
              s.cell(row=i, column=3).value = 'Core i3'
              print(span.text)
              wb.save('ST.xlsx')
              driver.close()
              break
            elif 'Core i9' in span.text:
                s.cell(row=i, column=3).value = 'Core i9'
                print(span.text)
                wb.save('ST.xlsx')
                driver.close()
                break
         except:
           print(" ST not correct ")
           s.cell(row=i, column=2).value = 'ST not correct'
           s.cell(row=i, column=3).value = 'ST not correct'
           s.cell(row=i, column=4).value = 'ST not correct'
           wb.save('ST.xlsx')
           driver.close()



         # for i in range(3, 35):
         # size = driver.find_elements(By.tagName("iframe"))
         # print(size)
         # core = WebDriverWait(driver, 20).until(
         #   EC.visibility_of_element_located((By.XPATH, "//<span>[contains(text(),‘Core’)]"))).text
         # driver.implicitly_wait(10)
         # core = driver.find_element(By.XPATH, "//div[text()='Core']").text
         # print(core)

zz = specs()
zz.pc_specs()
